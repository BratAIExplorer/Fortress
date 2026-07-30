import os
import sys
import time
import json
import logging
import io
import datetime
import pytz
import requests
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor, as_completed
from dotenv import load_dotenv
import threading
from zerodha_client import ZerodhaClient

# Global Broker variables
zerodha_client = None

# Define IST timezone
IST = pytz.timezone('Asia/Kolkata')

# Configure Logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(name)s: %(message)s',
    handlers=[
        logging.FileHandler("macd_bot.log", encoding="utf-8"),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger("MACDBot")

# File Paths
STATE_FILE = "macd_bot_state.json"
EXCEL_FILE = "MACD_Signals.xlsx"
CACHE_FILE = "nifty500_cache.csv"

# Load environment variables
load_dotenv()
TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
TELEGRAM_ADMIN_ID = os.getenv("TELEGRAM_ADMIN_ID")
FORTRESS_API_URL = os.getenv("FORTRESS_API_URL")  # e.g. https://fortressintelligence.space/api/analysis/momentum-signals
FORTRESS_CRON_SECRET = os.getenv("FORTRESS_CRON_SECRET")

if not TELEGRAM_BOT_TOKEN or not TELEGRAM_ADMIN_ID:
    logger.error("Missing TELEGRAM_BOT_TOKEN or TELEGRAM_ADMIN_ID in .env file!")
    sys.exit(1)

# Helper: Format Rupees
def format_rs(val):
    if pd.isna(val) or val is None:
        return "N/A"
    return f"₹{val:,.2f}"

# Helper: Format Percentage
def format_pct(val):
    if pd.isna(val) or val is None:
        return "0.00%"
    return f"{val:+.2f}%"

# Send Telegram Message
def send_telegram_message(text, reply_markup=None):
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
    payload = {
        "chat_id": TELEGRAM_ADMIN_ID,
        "text": text,
        "parse_mode": "Markdown"
    }
    if reply_markup:
        payload["reply_markup"] = reply_markup
    try:
        r = requests.post(url, json=payload, timeout=10)
        if r.status_code != 200:
            logger.error(f"Failed to send Telegram message: {r.text}")
            return False
        return True
    except Exception as e:
        logger.error(f"Error sending Telegram message: {e}")
        return False

# Load State File
def load_state():
    if os.path.exists(STATE_FILE):
        try:
            with open(STATE_FILE, "r") as f:
                return json.load(f)
        except Exception as e:
            logger.error(f"Error loading state file: {e}")
    return {"alerted_crossovers": {}}

# Save State File
def save_state(state):
    try:
        with open(STATE_FILE, "w") as f:
            json.dump(state, f, indent=4)
    except Exception as e:
        logger.error(f"Error saving state file: {e}")

# Fetch Nifty 500 Stock List
def get_nifty500_symbols():
    """
    Downloads Nifty 500 symbol list from NSE India.
    Falls back to cached CSV, and then to a default hardcoded list if all fail.
    """
    url = "https://archives.nseindia.com/content/indices/ind_nifty500list.csv"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }
    
    logger.info("Fetching Nifty 500 stock list from NSE...")
    try:
        r = requests.get(url, headers=headers, timeout=15)
        if r.status_code == 200:
            df = pd.read_csv(io.StringIO(r.text))
            if 'Symbol' in df.columns:
                symbols = df['Symbol'].dropna().unique().tolist()
                # Save to cache
                df.to_csv(CACHE_FILE, index=False)
                logger.info(f"Successfully loaded {len(symbols)} symbols from NSE and cached them.")
                return symbols
    except Exception as e:
        logger.warning(f"Failed to fetch Nifty 500 from NSE: {e}. Trying cached file.")
        
    # Check cache
    if os.path.exists(CACHE_FILE):
        try:
            df = pd.read_csv(CACHE_FILE)
            if 'Symbol' in df.columns:
                symbols = df['Symbol'].dropna().unique().tolist()
                logger.info(f"Successfully loaded {len(symbols)} symbols from local cache.")
                return symbols
        except Exception as e:
            logger.error(f"Failed to load Nifty 500 from cache: {e}")
            
    # Fallback to hardcoded list if both fail
    logger.warning("Using hardcoded fallback list of top Nifty stocks.")
    return [
        "RELIANCE", "TCS", "HDFCBANK", "INFY", "ICICIBANK", "HINDUNILVR", "ITC", "SBIN", "BHARTIARTL", "KOTAKBANK",
        "LT", "AXISBANK", "HCLTECH", "ASIANPAINT", "MARUTI", "SUNPHARMA", "TITAN", "BAJFINANCE", "ULTRACEMCO", "NTPC",
        "TATAMOTORS", "POWERGRID", "JSWSTEEL", "ADANIENT", "M&M", "COALINDIA", "WIPRO", "BAJAJFINSV", "NESTLEIND", "TATASTEEL",
        "IOC", "GRASIM", "TECHM", "BRITANNIA", "ADANIPORTS", "HINDALCO", "DRREDDY", "CIPLA", "TATACONSUM", "APOLLOHOSP",
        "EICHERMOT", "DIVISLAB", "BAJAJ-AUTO", "INDUSINDBK", "BPCL", "UPL", "HEROMOTOCO", "DABUR", "PIDILITIND", "GODREJCP"
    ]

# Fetch Live LTP (CMP) using Yahoo Finance Chart API (Single Ticker)
def fetch_live_ltp(symbol):
    """
    Fetches the live LTP (Last Traded Price) from Yahoo Finance chart API.
    Returns tuple of (symbol, ltp) or (symbol, None) if failed.
    """
    yf_symbol = f"{symbol}.NS"
    url = f"https://query1.finance.yahoo.com/v8/finance/chart/{yf_symbol}?interval=1m&range=1d"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Referer": "https://finance.yahoo.com/"
    }
    try:
        r = requests.get(url, headers=headers, timeout=10)
        if r.status_code == 200:
            data = r.json()
            meta = data.get("chart", {}).get("result", [{}])[0].get("meta", {})
            ltp = meta.get("regularMarketPrice")
            if ltp:
                return symbol, float(ltp)
    except Exception as e:
        logger.debug(f"Failed to fetch live LTP for {symbol}: {e}")
    return symbol, None

# Fetch Live LTP in Parallel
def fetch_all_live_prices(symbols):
    logger.info(f"Fetching live LTP for {len(symbols)} symbols in parallel...")
    prices = {}
    with ThreadPoolExecutor(max_workers=30) as executor:
        futures = {executor.submit(fetch_live_ltp, sym): sym for sym in symbols}
        for future in as_completed(futures):
            sym, ltp = future.result()
            if ltp is not None:
                prices[sym] = ltp
    logger.info(f"Successfully fetched live LTP for {len(prices)}/{len(symbols)} symbols.")
    return prices

# Download Historical Daily Candles in Chunks
def download_historical_data(symbols):
    """
    Downloads 1 year of daily historical Close price series for the symbols in chunks.
    Returns a dictionary of symbol -> pandas Series of Close prices.
    """
    import yfinance as yf
    
    chunk_size = 100
    all_series = {}
    
    logger.info(f"Downloading historical daily data for {len(symbols)} symbols in chunks of {chunk_size}...")
    
    for i in range(0, len(symbols), chunk_size):
        chunk = symbols[i : i + chunk_size]
        yf_symbols = [f"{sym}.NS" for sym in chunk]
        
        try:
            # Download Close prices only to save bandwidth and speed up
            data = yf.download(
                tickers=yf_symbols,
                period="1y",
                interval="1d",
                group_by="ticker",
                progress=False
            )
            
            # Extract the Close series for each symbol in chunk
            for sym in chunk:
                yf_sym = f"{sym}.NS"
                if yf_sym in data.columns.levels[0] if isinstance(data.columns, pd.MultiIndex) else yf_sym == data.name if isinstance(data, pd.Series) else False:
                    # MultiIndex columns case
                    close_series = data[yf_sym]['Close']
                    if not close_series.empty:
                        all_series[sym] = close_series
                elif yf_sym in data.columns:
                    # Flat columns case
                    close_series = data[yf_sym]
                    if not close_series.empty:
                        all_series[sym] = close_series
                        
        except Exception as e:
            logger.error(f"Error downloading chunk {i//chunk_size + 1}: {e}")
        
        time.sleep(0.5) # small delay to avoid rate limiting
        
    logger.info(f"Successfully downloaded history for {len(all_series)}/{len(symbols)} symbols.")
    return all_series

# Download Historical Weekly Candles in Chunks
def download_historical_weekly_data(symbols):
    """
    Downloads 5 years of weekly historical Close price series for the symbols in chunks.
    Returns a dictionary of symbol -> pandas Series of Close prices.
    """
    import yfinance as yf
    
    chunk_size = 100
    all_series = {}
    
    logger.info(f"Downloading historical weekly data for {len(symbols)} symbols in chunks of {chunk_size}...")
    
    for i in range(0, len(symbols), chunk_size):
        chunk = symbols[i : i + chunk_size]
        yf_symbols = [f"{sym}.NS" for sym in chunk]
        
        try:
            data = yf.download(
                tickers=yf_symbols,
                period="5y",
                interval="1wk",
                group_by="ticker",
                progress=False
            )
            
            # Extract the Close series for each symbol in chunk
            for sym in chunk:
                yf_sym = f"{sym}.NS"
                if yf_sym in data.columns.levels[0] if isinstance(data.columns, pd.MultiIndex) else yf_sym == data.name if isinstance(data, pd.Series) else False:
                    close_series = data[yf_sym]['Close']
                    if not close_series.empty:
                        all_series[sym] = close_series
                elif yf_sym in data.columns:
                    close_series = data[yf_sym]
                    if not close_series.empty:
                        all_series[sym] = close_series
                        
        except Exception as e:
            logger.error(f"Error downloading weekly chunk {i//chunk_size + 1}: {e}")
        
        time.sleep(0.5) # small delay to avoid rate limiting
        
    logger.info(f"Successfully downloaded weekly history for {len(all_series)}/{len(symbols)} symbols.")
    return all_series

# Calculate indicators, targets, SL and Position Sizing
def analyze_stock(symbol, close_series, ltp, timeframe="daily"):
    """
    Appends today's live LTP as the latest candle, calculates MACD & EMAs,
    checks for crossovers and drops, and computes targets, SL, and risk/reward.
    """
    # Ensure close_series is a pandas Series (squeeze if it is a DataFrame)
    if isinstance(close_series, pd.DataFrame):
        close_series = close_series.squeeze()
        
    # Make a copy of close_series and convert index to datetime without timezone
    close_series = close_series.copy()
    close_series.index = pd.to_datetime(close_series.index).tz_localize(None)
    
    if close_series.empty:
        return None

    if timeframe == "daily":
        # Get the latest date in the index BEFORE calling dropna()
        latest_date_in_db = close_series.index[-1].date()
        
        # Get today's date in IST
        today_ist = datetime.datetime.now(IST).date()
        
        # Check if today is a weekday (Monday-Friday)
        is_weekday = today_ist.weekday() < 5
        
        # Check if we are past market open (9:15 AM IST)
        now_ist_time = datetime.datetime.now(IST).time()
        is_after_market_open = now_ist_time >= datetime.time(9, 15)
        
        # If today is a weekday and we are past market open, the latest trading day is today.
        # Otherwise, the latest trading day is the latest date that yfinance returned.
        if is_weekday and is_after_market_open:
            target_date = today_ist
        else:
            target_date = latest_date_in_db
    else:  # timeframe == "weekly"
        today_ist = datetime.datetime.now(IST).date()
        # Monday of the current week
        monday_of_current_week = today_ist - datetime.timedelta(days=today_ist.weekday())
        
        # Check if we are past Monday 9:15 AM IST
        now_ist_time = datetime.datetime.now(IST).time()
        is_after_monday_open = True
        if today_ist.weekday() == 0:  # If today is Monday
            is_after_monday_open = now_ist_time >= datetime.time(9, 15)
        elif today_ist.weekday() > 0: # Tuesday-Sunday
            is_after_monday_open = True
        else:
            is_after_monday_open = False
            
        if is_after_monday_open:
            target_date = monday_of_current_week
        else:
            target_date = monday_of_current_week - datetime.timedelta(days=7)
            
    target_dt = pd.to_datetime(target_date)
    
    # Update or append the target date with the live LTP
    close_series[target_dt] = ltp
    
    # Clean the close series (drop any duplicate date index or other NaNs)
    close_series = close_series.dropna()
    close_series = close_series[~close_series.index.duplicated(keep='last')]
    close_series = close_series.sort_index()
    
    if len(close_series) < 200:
        # We need at least 200 days of history for 200 EMA
        return None
        
    # 3. Calculate MACD (12, 26, 9)
    ema12 = close_series.ewm(span=12, adjust=False).mean()
    ema26 = close_series.ewm(span=26, adjust=False).mean()
    macd_line = ema12 - ema26
    signal_line = macd_line.ewm(span=9, adjust=False).mean()
    
    # 4. Calculate EMAs (9, 20, 50, 100, 150, 200)
    ema9 = close_series.ewm(span=9, adjust=False).mean().iloc[-1]
    ema20 = close_series.ewm(span=20, adjust=False).mean().iloc[-1]
    ema50 = close_series.ewm(span=50, adjust=False).mean().iloc[-1]
    ema100 = close_series.ewm(span=100, adjust=False).mean().iloc[-1]
    ema150 = close_series.ewm(span=150, adjust=False).mean().iloc[-1]
    ema200 = close_series.ewm(span=200, adjust=False).mean().iloc[-1]
    
    # 5. Check Bullish Crossover & Recency (strictly 0 or 1 trading candles ago)
    macd_vals = macd_line.tolist()
    sig_vals = signal_line.tolist()
    dates = close_series.index.tolist()
    
    # Check latest crossover (index -1)
    crossover_today = (macd_vals[-1] > sig_vals[-1]) and (macd_vals[-2] <= sig_vals[-2])
    
    # Check previous crossover (index -2)
    crossover_yesterday = (macd_vals[-2] > sig_vals[-2]) and (macd_vals[-3] <= sig_vals[-3])
    
    # Current trend must be bullish (MACD above Signal today)
    is_bullish_today = macd_vals[-1] > sig_vals[-1]
    
    if not is_bullish_today:
        # Bearish: Drop immediately
        return {"action": "drop"}
        
    crossover_date = None
    days_since = None
    
    if crossover_today:
        crossover_date = dates[-1].strftime('%Y-%m-%d')
        days_since = 0
    elif crossover_yesterday:
        crossover_date = dates[-2].strftime('%Y-%m-%d')
        days_since = 1
    else:
        # Not within the 2-candle recency filter
        return None
        
    # Trend Filter: Ensure stock is in a primary uptrend (Price > 200 EMA and 50 EMA > 200 EMA)
    if ltp <= ema200 or ema50 <= ema200:
        return None
        
    # 6. Position Sizing
    capital = 25000
    qty = int(np.floor(capital / ltp))
    if qty <= 0:
        return None # Stock too expensive for 25k capital
    invested = qty * ltp
    
    # 7. EMA Targets & SL Calculations
    emas = {
        "EMA 20": float(ema20),
        "EMA 50": float(ema50),
        "EMA 100": float(ema100),
        "EMA 150": float(ema150),
        "EMA 200": float(ema200)
    }
    
    # Any EMA above CMP is a target candidate
    targets = {k: v for k, v in emas.items() if v > ltp}
    # Any EMA below CMP is a SL candidate
    supports = {k: v for k, v in emas.items() if v < ltp}
    
    # Determine First Target (nearest above CMP)
    if targets:
        first_target_ema_key = min(targets, key=targets.get)
        first_target_price = round(targets[first_target_ema_key], 2)
        first_target_ema = f"Weekly {first_target_ema_key}" if timeframe == "weekly" else first_target_ema_key
    else:
        first_target_ema = "5% Default (Blue Sky)"
        first_target_price = round(ltp * 1.05, 2)
        
    # Determine Final Target (topmost above CMP)
    if targets:
        final_target_ema_key = max(targets, key=targets.get)
        final_target_price = round(targets[final_target_ema_key], 2)
        final_target_ema = f"Weekly {final_target_ema_key}" if timeframe == "weekly" else final_target_ema_key
    else:
        final_target_ema = "15% Default (Blue Sky)"
        final_target_price = round(ltp * 1.15, 2)
        
    # Determine Stop Loss (nearest below CMP)
    if supports:
        sl_ema_key = max(supports, key=supports.get)
        sl_price = round(supports[sl_ema_key], 2)
        sl_ema = f"Weekly {sl_ema_key}" if timeframe == "weekly" else sl_ema_key
    else:
        sl_ema = "5% Default (SL)"
        sl_price = round(ltp * 0.95, 2)
        
    # Financial metrics
    profit_t1 = (first_target_price - ltp) * qty
    profit_final = (final_target_price - ltp) * qty
    risk_amount = (ltp - sl_price) * qty
    
    return {
        "action": "active",
        "data": {
            "Timeframe": "Daily" if timeframe == "daily" else "Weekly",
            "Symbol": symbol,
            "CMP": round(ltp, 2),
            "Crossover Date": crossover_date,
            "Days Since Crossover": days_since,
            "Quantity": qty,
            "Invested Amount": round(invested, 2),
            "First Target Price": first_target_price,
            "First Target EMA": first_target_ema,
            "Profit T1": round(profit_t1, 2),
            "Final Target Price": final_target_price,
            "Final Target EMA": final_target_ema,
            "Profit Final": round(profit_final, 2),
            "Stop Loss Price": sl_price,
            "Stop Loss EMA": sl_ema,
            "Risk Amount": round(risk_amount, 2)
        }
    }

# Update Excel Workbook
def push_signals_to_fortress(active_list):
    """
    Mirrors the 'Currently Active' sheet to the Fortress Momentum Radar tab
    (read-only display for other users; never places or influences orders).
    Best-effort only — never breaks the scan loop if Fortress is unreachable.
    """
    if not FORTRESS_API_URL or not FORTRESS_CRON_SECRET:
        return
    try:
        payload = {
            "signals": [
                {
                    "timeframe": item.get("Timeframe"),
                    "symbol": item.get("Symbol"),
                    "cmp": item.get("CMP"),
                    "crossoverDate": str(item.get("Crossover Date")),
                    "daysSinceCrossover": item.get("Days Since Crossover"),
                    "quantity": item.get("Quantity"),
                    "investedAmount": item.get("Invested Amount"),
                    "firstTargetPrice": item.get("First Target Price"),
                    "firstTargetEma": item.get("First Target EMA"),
                    "finalTargetPrice": item.get("Final Target Price"),
                    "finalTargetEma": item.get("Final Target EMA"),
                    "stopLossPrice": item.get("Stop Loss Price"),
                    "stopLossEma": item.get("Stop Loss EMA"),
                    "riskAmount": item.get("Risk Amount"),
                }
                for item in active_list
            ]
        }
        requests.post(
            FORTRESS_API_URL,
            json=payload,
            headers={"x-cron-secret": FORTRESS_CRON_SECRET},
            timeout=10,
        )
    except Exception as e:
        logger.warning(f"Fortress Momentum Radar push failed (non-blocking): {e}")


def update_excel_sheets(active_list, new_crossovers):
    """
    Writes active signals to 'Currently Active' sheet (overwriting).
    Appends new crossovers to 'Permanent Log' sheet (safely).
    """
    active_df = pd.DataFrame([item for item in active_list])
    
    # We will try to write to EXCEL_FILE. If permission is denied, we will write to EXCEL_FILE_LOCKED.xlsx
    # and send a warning.
    target_file = EXCEL_FILE
    use_fallback = False
    
    # Check if we can write to the file
    if os.path.exists(EXCEL_FILE):
        try:
            # Attempt to open the file in write mode to see if it is locked
            with open(EXCEL_FILE, 'a'):
                pass
        except OSError:
            # File is locked!
            target_file = "MACD_Signals_LOCKED.xlsx"
            use_fallback = True
            logger.warning(f"Excel file {EXCEL_FILE} is locked. Writing to {target_file} instead.")
            send_telegram_message("⚠️ *Warning*: Excel file `MACD_Signals.xlsx` is open or locked. Latest signals saved to `MACD_Signals_LOCKED.xlsx` instead. Please close `MACD_Signals.xlsx` to resume standard logging.")

    # Self-Healing Format Upgrade: Check if existing file lacks "Timeframe" column
    if os.path.exists(target_file):
        need_recreate = False
        try:
            import gc
            book = load_workbook(target_file, read_only=True)
            if 'Permanent Log' in book.sheetnames:
                sheet = book['Permanent Log']
                headers = []
                for row in sheet.iter_rows(max_row=1, values_only=True):
                    headers = list(row)
                    break
                if "Timeframe" not in headers:
                    need_recreate = True
            book.close()
            del book
            gc.collect()
        except Exception as e:
            logger.error(f"Error checking column structure of {target_file}: {e}")
            
        if need_recreate:
            logger.info("Excel workbook has old schema (missing Timeframe column). Migrating...")
            backup_file = target_file.replace(".xlsx", f"_backup_{int(time.time())}.xlsx")
            try:
                time.sleep(0.2)
                os.rename(target_file, backup_file)
                logger.info(f"Backed up old Excel workbook to {backup_file}")
            except Exception as ex:
                logger.error(f"Failed to backup old Excel file {target_file}: {ex}")

    # Columns definitions
    active_cols = [
        "Timeframe", "Symbol", "CMP", "Crossover Date", "Days Since Crossover",
        "Quantity", "Invested Amount", "First Target Price", "First Target EMA",
        "Final Target Price", "Final Target EMA", "Stop Loss Price", "Stop Loss EMA", "Risk Amount"
    ]
    log_columns = [
        "Timestamp", "Timeframe", "Symbol", "CMP", "Crossover Date", "Days Since Crossover",
        "Quantity", "Invested Amount", "First Target Price", "First Target EMA",
        "Final Target Price", "Final Target EMA", "Stop Loss Price", "Stop Loss EMA", "Risk Amount"
    ]

    # 1. Handle Entire Workbook creation if not exists
    if not os.path.exists(target_file):
        try:
            with pd.ExcelWriter(target_file, engine='openpyxl') as writer:
                if not active_df.empty:
                    # ensure all columns exist
                    for c in active_cols:
                        if c not in active_df.columns:
                            active_df[c] = None
                    active_df[active_cols].to_excel(writer, sheet_name='Currently Active', index=False)
                else:
                    pd.DataFrame(columns=active_cols).to_excel(writer, sheet_name='Currently Active', index=False)
                    
                pd.DataFrame(columns=log_columns).to_excel(writer, sheet_name='Permanent Log', index=False)
        except Exception as e:
            logger.error(f"Error creating excel workbook {target_file}: {e}")
            return
            
    # 2. Overwrite 'Currently Active' sheet
    try:
        book = load_workbook(target_file)
        if 'Currently Active' in book.sheetnames:
            book.remove(book['Currently Active'])
        book.save(target_file)
        book.close()
        
        with pd.ExcelWriter(target_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            if not active_df.empty:
                for c in active_cols:
                    if c not in active_df.columns:
                        active_df[c] = None
                active_df[active_cols].to_excel(writer, sheet_name='Currently Active', index=False)
            else:
                pd.DataFrame(columns=active_cols).to_excel(writer, sheet_name='Currently Active', index=False)
    except Exception as e:
        logger.error(f"Error overwriting Currently Active sheet in {target_file}: {e}")

    # 3. Append to 'Permanent Log' sheet
    if new_crossovers:
        try:
            timestamp = datetime.datetime.now(IST).strftime('%Y-%m-%d %H:%M:%S')
            log_entries = []
            for item in new_crossovers:
                entry = {
                    "Timestamp": timestamp,
                    "Timeframe": item["Timeframe"],
                    "Symbol": item["Symbol"],
                    "CMP": item["CMP"],
                    "Crossover Date": item["Crossover Date"],
                    "Days Since Crossover": item["Days Since Crossover"],
                    "Quantity": item["Quantity"],
                    "Invested Amount": item["Invested Amount"],
                    "First Target Price": item["First Target Price"],
                    "First Target EMA": item["First Target EMA"],
                    "Final Target Price": item["Final Target Price"],
                    "Final Target EMA": item["Final Target EMA"],
                    "Stop Loss Price": item["Stop Loss Price"],
                    "Stop Loss EMA": item["Stop Loss EMA"],
                    "Risk Amount": item["Risk Amount"]
                }
                log_entries.append(entry)
                
            log_df = pd.DataFrame(log_entries)
            
            book = load_workbook(target_file)
            sheet = book['Permanent Log']
            start_row = sheet.max_row + 1
            
            for r_idx, row in enumerate(log_df[log_columns].values, start=start_row):
                for c_idx, val in enumerate(row, start=1):
                    sheet.cell(row=r_idx, column=c_idx, value=val)
                    
            book.save(target_file)
            book.close()
            logger.info(f"Successfully appended {len(new_crossovers)} entries to Permanent Log in {target_file}.")
        except Exception as e:
            logger.error(f"Error appending to Permanent Log in {target_file}: {e}")

# Monitor tracked positions for Target/SL hit alerts
def check_tracked_positions(state, live_prices):
    tracked_positions = state.setdefault("tracked_positions", {})
    if not tracked_positions:
        return
        
    logger.info(f"Checking targets/SL for {len(tracked_positions)} tracked positions...")
    updated = False
    
    for sym in list(tracked_positions.keys()):
        pos = tracked_positions[sym]
        ltp = live_prices.get(sym)
        if not ltp:
            continue
            
        t1 = pos.get("t1", 0.0)
        t2 = pos.get("t2", 0.0)
        sl = pos.get("sl", 0.0)
        qty = pos.get("qty", 0)
        
        # 1. Check Stop Loss
        if ltp <= sl and not pos.get("sl_hit", False):
            pos["sl_hit"] = True
            updated = True
            msg = (
                f"⚠️ *Stop Loss Hit Alert!*\n\n"
                f"● *Stock*: **{sym}**\n"
                f"● *Current Price*: {format_rs(ltp)}\n"
                f"● *Stop Loss Level*: {format_rs(sl)}\n"
                f"● *Quantity*: {qty}\n\n"
                f"💡 *Action*: Click below to sell and exit now:"
            )
            reply_markup = {
                "inline_keyboard": [
                    [{"text": f"🛑 Sell {qty} Qty", "callback_data": f"sell:{sym}:{qty}"}]
                ]
            }
            send_telegram_message(msg, reply_markup=reply_markup)
            
        # 2. Check Target 1
        elif ltp >= t1 and not pos.get("t1_hit", False):
            pos["t1_hit"] = True
            updated = True
            msg = (
                f"🎯 *Target 1 Met Alert!*\n\n"
                f"● *Stock*: **{sym}**\n"
                f"● *Current Price*: {format_rs(ltp)}\n"
                f"● *Target 1 Level*: {format_rs(t1)}\n"
                f"● *Quantity*: {qty}\n\n"
                f"💡 *Action*: Click below to book profit now:"
            )
            reply_markup = {
                "inline_keyboard": [
                    [{"text": f"🛑 Sell {qty} Qty", "callback_data": f"sell:{sym}:{qty}"}]
                ]
            }
            send_telegram_message(msg, reply_markup=reply_markup)
            
        # 3. Check Target 2 (Final Target)
        elif ltp >= t2 and not pos.get("t2_hit", False):
            pos["t2_hit"] = True
            updated = True
            msg = (
                f"🎯 *Final Target Met Alert!*\n\n"
                f"● *Stock*: **{sym}**\n"
                f"● *Current Price*: {format_rs(ltp)}\n"
                f"● *Final Target Level*: {format_rs(t2)}\n"
                f"● *Quantity*: {qty}\n\n"
                f"💡 *Action*: Click below to exit now:"
            )
            reply_markup = {
                "inline_keyboard": [
                    [{"text": f"🛑 Sell {qty} Qty", "callback_data": f"sell:{sym}:{qty}"}]
                ]
            }
            send_telegram_message(msg, reply_markup=reply_markup)
            
    if updated:
        save_state(state)

# Log daily trades in state file
def log_daily_trade(action, symbol, qty, price_or_pnl):
    try:
        state = load_state()
        today_str = datetime.datetime.now(IST).strftime('%Y-%m-%d')
        daily_trades = state.setdefault("daily_trades", {})
        today_trades = daily_trades.setdefault(today_str, {"buys": [], "sells": []})
        
        if action.lower() == "buy":
            # Check if trade already logged today to prevent duplicate logs on re-clicks
            if not any(b["symbol"] == symbol and b["qty"] == qty for b in today_trades["buys"]):
                today_trades["buys"].append({
                    "symbol": symbol,
                    "qty": qty,
                    "invested": round(price_or_pnl, 2),
                    "time": datetime.datetime.now(IST).strftime('%H:%M:%S')
                })
        elif action.lower() == "sell":
            if not any(s["symbol"] == symbol and s["qty"] == qty for s in today_trades["sells"]):
                today_trades["sells"].append({
                    "symbol": symbol,
                    "qty": qty,
                    "pnl": round(price_or_pnl, 2),
                    "time": datetime.datetime.now(IST).strftime('%H:%M:%S')
                })
                # Update cumulative P&L
                cum_pnl = state.setdefault("cumulative_pnl", 0.0)
                state["cumulative_pnl"] = round(cum_pnl + price_or_pnl, 2)
        save_state(state)
        logger.info(f"Logged daily trade: {action.upper()} {symbol} (Qty: {qty})")
    except Exception as e:
        logger.error(f"Error logging daily trade: {e}")

# Compile and send the daily summary report to Telegram
def send_daily_summary(state):
    global zerodha_client
    now = datetime.datetime.now(IST)
    today_str = now.strftime('%Y-%m-%d')
    
    logger.info("Generating Daily P&L Summary Report...")
    
    # Get today's trades from state
    daily_trades = state.get("daily_trades", {}).get(today_str, {"buys": [], "sells": []})
    buys = daily_trades.get("buys", [])
    sells = daily_trades.get("sells", [])
    
    # Calculate Total Profit Booked
    total_pnl = sum(s["pnl"] for s in sells)
    cum_pnl = state.get("cumulative_pnl", 0.0)
    
    msg = f"📊 *Daily Summary - {now.strftime('%b %d, %Y')}*\n\n"
    
    # 1. Booked Profits Section
    if sells:
        msg += "💰 *Profit Booked Today:*\n"
        for s in sells:
            msg += f"• **{s['symbol']}**: {s['qty']} Qty | Profit: {format_rs(s['pnl'])}\n"
        msg += f"↳ **Total Realized P&L: {format_rs(total_pnl)}**\n"
        msg += f"🏆 **Cumulative Realized P&L: {format_rs(cum_pnl)}**\n\n"
    else:
        msg += "💰 *Profit Booked Today:* None\n"
        msg += f"🏆 **Cumulative Realized P&L: {format_rs(cum_pnl)}**\n\n"
        
    # 2. New Buys Section
    if buys:
        msg += "🛒 *New Positions Entered Today:*\n"
        for b in buys:
            msg += f"• **{b['symbol']}**: {b['qty']} Qty | Invested: {format_rs(b['invested'])}\n"
        msg += "\n"
    else:
        msg += "🛒 *New Positions Entered Today:* None\n\n"
        
    # 3. Portfolio Health Snapshot (Holdings & Positions)
    portfolio_details = ""
    total_holdings_value = 0.0
    total_unrealized_pnl = 0.0
    
    if zerodha_client and zerodha_client.logged_in:
        try:
            # Holdings
            holdings = zerodha_client.kite.holdings()
            for h in holdings:
                symbol = h["tradingsymbol"]
                qty = h["quantity"]
                pnl = h.get("pnl", 0.0)
                last_price = h.get("last_price", 0.0)
                value = qty * last_price
                
                if qty > 0:
                    total_holdings_value += value
                    total_unrealized_pnl += pnl
                    portfolio_details += f"• **{symbol}** (Holding): {qty} Qty | P&L: {format_rs(pnl)}\n"
                    
            # Positions
            positions_data = zerodha_client.kite.positions()
            net_positions = positions_data.get("net", [])
            for pos in net_positions:
                symbol = pos["tradingsymbol"]
                qty = pos["quantity"]
                pnl = pos.get("unrealised", 0.0)
                last_price = pos.get("last_price", 0.0)
                value = qty * last_price
                
                if qty > 0:
                    total_holdings_value += value
                    total_unrealized_pnl += pnl
                    portfolio_details += f"• **{symbol}** (Position): {qty} Qty | P&L: {format_rs(pnl)}\n"
        except Exception as ex:
            logger.error(f"Error fetching portfolio snapshot for daily summary: {ex}")
            portfolio_details = "⚠️ *Failed to fetch live portfolio snapshot.*"
            
    if portfolio_details:
        msg += "📈 *Active Portfolio Snapshot:*\n"
        msg += portfolio_details
        msg += f"↳ **Total Portfolio Value: {format_rs(total_holdings_value)}**\n"
        msg += f"↳ **Unrealized MTM P&L: {format_rs(total_unrealized_pnl)}**\n"
    else:
        msg += "📈 *Active Portfolio Snapshot:* No open trades.\n"
        
    # Send message
    success = send_telegram_message(msg)
    if success:
        state["last_summary_date"] = today_str
        save_state(state)
        logger.info("Daily Summary report sent successfully.")
    else:
        logger.error("Failed to send Daily Summary report.")

# Scan current holdings and positions for ₹300 profit target
def check_and_autosell_holdings():
    global zerodha_client
    if not zerodha_client or not zerodha_client.logged_in:
        logger.info("Auto-sell check skipped: Zerodha client is not logged in.")
        return

    logger.info("Scanning current holdings & positions for Rs. 300 profit target...")
    processed_sells = set()
    
    try:
        # 1. Scan Holdings (multi-day settled positions)
        holdings = zerodha_client.kite.holdings()
        for holding in holdings:
            symbol = holding["tradingsymbol"]
            qty = holding["quantity"]
            pnl = holding.get("pnl", 0.0)
            
            if qty > 0 and pnl >= 300.0 and symbol not in processed_sells:
                logger.info(f"🤖 Auto-sell trigger (Holding): {symbol} has Rs. {pnl:.2f} profit (>= Rs. 300). Placing market sell order...")
                processed_sells.add(symbol)
                
                success, msg = execute_broker_order("sell", symbol, qty)
                if success:
                    log_daily_trade("sell", symbol, qty, pnl)
                    send_telegram_message(
                        f"🤖 *Auto-Sell Triggered (Holding)!*\n\n"
                        f"● *Stock*: **{symbol}**\n"
                        f"● *Quantity*: {qty}\n"
                        f"● *Profit Booked*: {format_rs(pnl)}\n"
                        f"● *Status*: Sold successfully on Zerodha."
                    )
                    # Clean up tracked positions state if tracked locally
                    state = load_state()
                    tracked_positions = state.setdefault("tracked_positions", {})
                    if symbol in tracked_positions:
                        del tracked_positions[symbol]
                        save_state(state)
                else:
                    send_telegram_message(f"❌ *Auto-Sell Failed* for holding {symbol}: `{msg}`")

        # 2. Scan Positions (same-day open positions)
        positions_data = zerodha_client.kite.positions()
        net_positions = positions_data.get("net", [])
        for pos in net_positions:
            symbol = pos["tradingsymbol"]
            qty = pos["quantity"]
            pnl = pos.get("unrealised", 0.0)
            
            if qty > 0 and pnl >= 300.0 and symbol not in processed_sells:
                logger.info(f"🤖 Auto-sell trigger (Position): {symbol} has Rs. {pnl:.2f} profit (>= Rs. 300). Placing market sell order...")
                processed_sells.add(symbol)
                
                success, msg = execute_broker_order("sell", symbol, qty)
                if success:
                    log_daily_trade("sell", symbol, qty, pnl)
                    send_telegram_message(
                        f"🤖 *Auto-Sell Triggered (Position)!*\n\n"
                        f"● *Stock*: **{symbol}**\n"
                        f"● *Quantity*: {qty}\n"
                        f"● *Profit Booked*: {format_rs(pnl)}\n"
                        f"● *Status*: Sold successfully on Zerodha."
                    )
                    # Clean up tracked positions state if tracked locally
                    state = load_state()
                    tracked_positions = state.setdefault("tracked_positions", {})
                    if symbol in tracked_positions:
                        del tracked_positions[symbol]
                        save_state(state)
                else:
                    send_telegram_message(f"❌ *Auto-Sell Failed* for position {symbol}: `{msg}`")
    except Exception as e:
        logger.error(f"Error checking holdings/positions for auto-sell: {e}")

# Run One Full Scan Cycle
def run_scan_cycle(state):
    logger.info("Starting a new scan cycle...")
    
    # Check session validity if the client is initialized and logged in
    global zerodha_client
    if zerodha_client and zerodha_client.logged_in:
        if not zerodha_client.validate_session():
            logger.warning("Zerodha session expired during scan cycle.")
            trigger_zerodha_login(force=False)
            
    # 1. Fetch symbols
    symbols = get_nifty500_symbols()
    if not symbols:
        logger.error("No Nifty 500 symbols found!")
        return
        
    # 2. Fetch live prices in parallel
    live_prices = fetch_all_live_prices(symbols)
    valid_symbols = list(live_prices.keys())
    
    # 3. Fetch historical daily series in chunks
    historical_data = download_historical_data(valid_symbols)
    
    # 4. Fetch historical weekly series in chunks
    historical_weekly_data = download_historical_weekly_data(valid_symbols)
    
    active_list = []
    new_crossovers = []
    
    # Migrate alerted_crossovers in state if needed
    alerted_crossovers = state.setdefault("alerted_crossovers", {})
    if alerted_crossovers and not any(k in alerted_crossovers for k in ["Daily", "Weekly"]):
        old_dict = alerted_crossovers.copy()
        alerted_crossovers.clear()
        alerted_crossovers["Daily"] = old_dict
        alerted_crossovers["Weekly"] = {}
    elif not alerted_crossovers:
        alerted_crossovers["Daily"] = {}
        alerted_crossovers["Weekly"] = {}
        
    alerted_daily = alerted_crossovers.setdefault("Daily", {})
    alerted_weekly = alerted_crossovers.setdefault("Weekly", {})
    
    # Daily Scan
    logger.info("Analyzing daily stock data...")
    for sym in valid_symbols:
        close_series = historical_data.get(sym)
        if close_series is None or close_series.empty:
            continue
            
        ltp = live_prices[sym]
        
        try:
            res = analyze_stock(sym, close_series, ltp, timeframe="daily")
            if res is None:
                continue
                
            # If stock crosses bearish, we clean up state so it can be re-triggered next time
            if res["action"] == "drop":
                if sym in alerted_daily:
                    del alerted_daily[sym]
                continue
                
            if res["action"] == "active":
                data = res["data"]
                active_list.append(data)
                
                # Check if this crossover is fresh and needs alerting
                crossover_date = data["Crossover Date"]
                last_alerted_date = alerted_daily.get(sym)
                
                if last_alerted_date != crossover_date:
                    # New fresh crossover!
                    new_crossovers.append(data)
                    alerted_daily[sym] = crossover_date
                    
        except Exception as e:
            logger.error(f"Error analyzing daily symbol {sym}: {e}")
            
    # Weekly Scan
    logger.info("Analyzing weekly stock data...")
    for sym in valid_symbols:
        close_series = historical_weekly_data.get(sym)
        if close_series is None or close_series.empty:
            continue
            
        ltp = live_prices[sym]
        
        try:
            res = analyze_stock(sym, close_series, ltp, timeframe="weekly")
            if res is None:
                continue
                
            # If stock crosses bearish, we clean up state so it can be re-triggered next time
            if res["action"] == "drop":
                if sym in alerted_weekly:
                    del alerted_weekly[sym]
                continue
                
            if res["action"] == "active":
                data = res["data"]
                active_list.append(data)
                
                # Check if this crossover is fresh and needs alerting
                crossover_date = data["Crossover Date"]
                last_alerted_date = alerted_weekly.get(sym)
                
                if last_alerted_date != crossover_date:
                    # New fresh crossover!
                    new_crossovers.append(data)
                    alerted_weekly[sym] = crossover_date
                    
        except Exception as e:
            logger.error(f"Error analyzing weekly symbol {sym}: {e}")
            
    # 5. Save state
    save_state(state)
    
    # 6. Update Excel Workbook
    update_excel_sheets(active_list, new_crossovers)

    # 6b. Mirror to Fortress Momentum Radar (best-effort, non-blocking)
    push_signals_to_fortress(active_list)

    # 7. Alert Telegram
    logger.info(f"Completed scan. Active: {len(active_list)}. New crossovers: {len(new_crossovers)}.")
    for item in new_crossovers:
        tf = item["Timeframe"]
        period_unit = "days" if tf == "Daily" else "weeks"
        msg = (
            f"📢 *New {tf} Bullish Signal Detected!*\n\n"
            f"● *Stock*: **{item['Symbol']}**\n"
            f"● *CMP*: {format_rs(item['CMP'])}\n"
            f"● *Crossover Date*: `{item['Crossover Date']}` ({item['Days Since Crossover']} {period_unit} ago)\n"
            f"● *Quantity to Buy*: **{item['Quantity']}** (Capital: ₹25,000)\n"
            f"● *Invested Amount*: {format_rs(item['Invested Amount'])}\n\n"
            f"🎯 *First Target*: **{format_rs(item['First Target Price'])}** ({item['First Target EMA']})\n"
            f"  ↳ Expected Profit: {format_rs(item['Profit T1'])}\n"
            f"🎯 *Final Target*: **{format_rs(item['Final Target Price'])}** ({item['Final Target EMA']})\n"
            f"  ↳ Expected Profit: {format_rs(item['Profit Final'])}\n"
            f"🛑 *Stop Loss*: **{format_rs(item['Stop Loss Price'])}** ({item['Stop Loss EMA']})\n"
            f"  ↳ Risk Amount: {format_rs(item['Risk Amount'])}\n\n"
            f"⚠️ *Disclaimer*: This is not financial advice. Past performance is not indicative of future results. Please consult with a financial advisor before trading.\n\n"
            f"💡 *Action*: If you buy it now at CMP of {format_rs(item['CMP'])}, these number of quantities ({item['Quantity']}), "
            f"the expected profit is {format_rs(item['Profit T1'])} (Target 1) or {format_rs(item['Profit Final'])} (Final Target) "
            f"with a stop loss of {format_rs(item['Stop Loss Price'])} basis the above conditions."
        )
        reply_markup = {
            "inline_keyboard": [
                [
                    {"text": f"🛒 Buy {item['Quantity']} Qty", "callback_data": f"buy:{item['Symbol']}:{item['Quantity']}:{item['First Target Price']:.2f}:{item['Final Target Price']:.2f}:{item['Stop Loss Price']:.2f}"},
                    {"text": f"🛑 Sell {item['Quantity']} Qty", "callback_data": f"sell:{item['Symbol']}:{item['Quantity']}"}
                ],
                [
                    {"text": "➖ 50%", "callback_data": f"qty_half:{item['Symbol']}:{item['Quantity']}:{item['First Target Price']:.2f}:{item['Final Target Price']:.2f}:{item['Stop Loss Price']:.2f}"},
                    {"text": "➕ Double", "callback_data": f"qty_double:{item['Symbol']}:{item['Quantity']}:{item['First Target Price']:.2f}:{item['Final Target Price']:.2f}:{item['Stop Loss Price']:.2f}"},
                    {"text": "✏️ Custom", "callback_data": f"qty_custom:{item['Symbol']}:{item['First Target Price']:.2f}:{item['Final Target Price']:.2f}:{item['Stop Loss Price']:.2f}"}
                ]
            ]
        }
        send_telegram_message(msg, reply_markup=reply_markup)
        time.sleep(1.0) # small sleep to avoid telegram rate limits

    # Check live prices for tracked positions (Targets / Stop Loss met alerts)
    check_tracked_positions(state, live_prices)

# Check if current time is within Indian market hours (9:15 - 15:30 IST, Mon-Fri)
def is_market_hours():
    now = datetime.datetime.now(IST)
    # Weekday 0-4 is Mon-Fri, 5-6 is Sat-Sun
    if now.weekday() >= 5:
        return False
        
    start_time = datetime.time(9, 15)
    end_time = datetime.time(15, 30)
    current_time = now.time()
    
    return start_time <= current_time <= end_time

# Trigger login sequence to prompt for Zerodha login on Telegram
def trigger_zerodha_login(force=False):
    global zerodha_client
    
    # Load state to check daily limit
    state = load_state()
    today_str = datetime.datetime.now(IST).strftime('%Y-%m-%d')
    
    if not force:
        last_login_prompt_date = state.get("last_login_prompt_date")
        if last_login_prompt_date == today_str:
            logger.info("A login prompt was already sent today. Skipping auto-prompt to avoid spam.")
            send_telegram_message(
                "⚠️ *Zerodha Session Expired*\n"
                "Auto-login prompt skipped to avoid duplicate alerts today.\n"
                "👉 Send `login` to receive the login link manually."
            )
            return False

    if not zerodha_client:
        return False

    login_url = zerodha_client.get_login_url()
    if login_url:
        msg = (
            "🔑 *Zerodha Login Required*\n\n"
            "Please click the link below to log in and authorize the app:\n"
            f"🔗 [Authorize Zerodha]({login_url})\n\n"
            "After logging in, you will be redirected to a page (e.g. `https://127.0.0.1:8000`).\n"
            "👉 *Copy the entire redirected URL* (or just the `request_token` from the address bar) and paste it here."
        )
        send_telegram_message(msg)
        
        # Save today's date as the last prompt date
        state["last_login_prompt_date"] = today_str
        save_state(state)
        return True
    else:
        send_telegram_message("❌ Failed to generate Zerodha login URL. Send `login` to try again.")
        return False

# Execute market order using Zerodha client
def execute_broker_order(action, symbol, qty):
    global zerodha_client
    if not zerodha_client or not zerodha_client.logged_in:
        return False, "Broker client is not logged in."
    
    exchange = "NSE"
    success, message = zerodha_client.place_order(symbol, exchange, qty, action.upper())
    return success, message

# Process callback query (button click)
def handle_callback_query(callback_query):
    query_id = callback_query["id"]
    chat_id = callback_query["message"]["chat"]["id"]
    message_id = callback_query["message"]["message_id"]
    data = callback_query.get("data", "")
    
    if str(chat_id) != str(TELEGRAM_ADMIN_ID):
        return
        
    if data.startswith("qty_half:") or data.startswith("qty_double:"):
        parts = data.split(":")
        action = parts[0]
        symbol = parts[1]
        qty = int(parts[2])
        t1 = float(parts[3]) if len(parts) > 3 else 0.0
        t2 = float(parts[4]) if len(parts) > 4 else 0.0
        sl = float(parts[5]) if len(parts) > 5 else 0.0
        
        if action == "qty_half":
            new_qty = max(1, qty // 2)
        else:
            new_qty = qty * 2
            
        reply_markup = {
            "inline_keyboard": [
                [
                    {"text": f"🛒 Buy {new_qty} Qty", "callback_data": f"buy:{symbol}:{new_qty}:{t1:.2f}:{t2:.2f}:{sl:.2f}"},
                    {"text": f"🛑 Sell {new_qty} Qty", "callback_data": f"sell:{symbol}:{new_qty}"}
                ],
                [
                    {"text": "➖ 50%", "callback_data": f"qty_half:{symbol}:{new_qty}:{t1:.2f}:{t2:.2f}:{sl:.2f}"},
                    {"text": "➕ Double", "callback_data": f"qty_double:{symbol}:{new_qty}:{t1:.2f}:{t2:.2f}:{sl:.2f}"},
                    {"text": "✏️ Custom", "callback_data": f"qty_custom:{symbol}:{t1:.2f}:{t2:.2f}:{sl:.2f}"}
                ]
            ]
        }
        
        requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/editMessageReplyMarkup",
            json={
                "chat_id": chat_id,
                "message_id": message_id,
                "reply_markup": reply_markup
            }
        )
        
        requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/answerCallbackQuery",
            json={"callback_query_id": query_id, "text": f"Quantity updated to {new_qty}."}
        )
        return
        
    elif data.startswith("qty_custom:"):
        parts = data.split(":")
        symbol = parts[1]
        t1 = float(parts[2]) if len(parts) > 2 else 0.0
        t2 = float(parts[3]) if len(parts) > 3 else 0.0
        sl = float(parts[4]) if len(parts) > 4 else 0.0
        
        state = load_state()
        state["waiting_for_custom_qty"] = {
            "symbol": symbol,
            "message_id": message_id,
            "t1": t1,
            "t2": t2,
            "sl": sl
        }
        save_state(state)
        
        requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/answerCallbackQuery",
            json={"callback_query_id": query_id, "text": "Waiting for custom quantity input..."}
        )
        
        send_telegram_message(f"✏️ *Custom Quantity for {symbol}*\n\nPlease reply by typing the custom quantity number (e.g. `15`), or type `cancel` to abort.")
        return
        
    elif data.startswith("buy:") or data.startswith("sell:"):
        parts = data.split(":")
        action = parts[0]
        symbol = parts[1]
        qty = int(parts[2])
        
        # Parse Targets and Stop Loss values if present
        t1 = float(parts[3]) if len(parts) > 3 else 0.0
        t2 = float(parts[4]) if len(parts) > 4 else 0.0
        sl = float(parts[5]) if len(parts) > 5 else 0.0
        
        # Acknowledge callback query
        requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/answerCallbackQuery",
            json={"callback_query_id": query_id, "text": f"Processing {action.upper()} order..."}
        )
        
        send_telegram_message(f"⏳ Placing market {action.upper()} order for {qty} shares of {symbol}...")
        
        success, err_message = execute_broker_order(action, symbol, qty)
        
        if success:
            original_text = callback_query["message"].get("text", "")
            new_text = original_text + f"\n\n[Executed: {action.upper()} {qty} Qty]"
            requests.post(
                f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/editMessageText",
                json={
                    "chat_id": chat_id,
                    "message_id": message_id,
                    "text": new_text
                }
            )
            
            # Manage tracked positions state and log daily trade
            state = load_state()
            if action == "buy":
                # Fetch current live price to log purchase price
                _, ltp = fetch_live_ltp(symbol)
                buy_price = ltp if ltp else 0.0
                invested = (buy_price * qty) if buy_price else 0.0
                
                tracked_positions = state.setdefault("tracked_positions", {})
                tracked_positions[symbol] = {
                    "symbol": symbol,
                    "qty": qty,
                    "buy_price": buy_price,
                    "t1": t1,
                    "t2": t2,
                    "sl": sl,
                    "t1_hit": False,
                    "t2_hit": False,
                    "sl_hit": False
                }
                save_state(state)
                log_daily_trade("buy", symbol, qty, invested)
            elif action == "sell":
                tracked_positions = state.setdefault("tracked_positions", {})
                pnl = 0.0
                if symbol in tracked_positions:
                    pos = tracked_positions[symbol]
                    _, ltp = fetch_live_ltp(symbol)
                    buy_price = pos.get("buy_price", 0.0)
                    if buy_price > 0 and ltp:
                        pnl = (ltp - buy_price) * qty
                    del tracked_positions[symbol]
                    save_state(state)
                log_daily_trade("sell", symbol, qty, pnl)
        else:
            send_telegram_message(f"❌ *Order Failed* for {symbol}: `{err_message}`")

# Process incoming text message (token/URL or 'login')
def handle_text_message(message):
    global zerodha_client
    chat_id = message["chat"]["id"]
    text = message.get("text", "").strip()
    
    if str(chat_id) != str(TELEGRAM_ADMIN_ID):
        return
        
    # Check if we are waiting for a custom quantity
    state = load_state()
    waiting = state.get("waiting_for_custom_qty")
    if waiting:
        if text.lower() == "cancel":
            del state["waiting_for_custom_qty"]
            save_state(state)
            send_telegram_message("Cancelled quantity update.")
            return
            
        try:
            new_qty = int(text)
            if new_qty <= 0:
                raise ValueError("Quantity must be positive.")
                
            symbol = waiting["symbol"]
            orig_msg_id = waiting["message_id"]
            t1 = waiting["t1"]
            t2 = waiting["t2"]
            sl = waiting["sl"]
            
            # Update the original message's keyboard with the new qty!
            reply_markup = {
                "inline_keyboard": [
                    [
                        {"text": f"🛒 Buy {new_qty} Qty", "callback_data": f"buy:{symbol}:{new_qty}:{t1:.2f}:{t2:.2f}:{sl:.2f}"},
                        {"text": f"🛑 Sell {new_qty} Qty", "callback_data": f"sell:{symbol}:{new_qty}"}
                    ],
                    [
                        {"text": "➖ 50%", "callback_data": f"qty_half:{symbol}:{new_qty}:{t1:.2f}:{t2:.2f}:{sl:.2f}"},
                        {"text": "➕ Double", "callback_data": f"qty_double:{symbol}:{new_qty}:{t1:.2f}:{t2:.2f}:{sl:.2f}"},
                        {"text": "✏️ Custom", "callback_data": f"qty_custom:{symbol}:{t1:.2f}:{t2:.2f}:{sl:.2f}"}
                    ]
                ]
            }
            
            # Send edit message reply markup request
            requests.post(
                f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/editMessageReplyMarkup",
                json={
                    "chat_id": chat_id,
                    "message_id": orig_msg_id,
                    "reply_markup": reply_markup
                }
            )
            
            send_telegram_message(f"✅ Quantity for {symbol} updated to **{new_qty}**.")
            
            # Clear waiting state
            del state["waiting_for_custom_qty"]
            save_state(state)
            return
            
        except ValueError:
            send_telegram_message("❌ Invalid quantity. Please send a valid positive number, or send `cancel` to abort.")
            return

    # Check if user pasted a token/URL or typed 'login'
    is_token_or_url = "request_token=" in text or (text.isalnum() and 20 <= len(text) <= 64)
    
    if is_token_or_url:
        send_telegram_message("⏳ Verifying token and generating Zerodha session...")
        access_token = zerodha_client.verify_request_token_and_login(text)
        if access_token:
            send_telegram_message("✅ *Zerodha Login Successful!* Session is active.")
            # Cache the session token to the state
            state = load_state()
            state["zerodha_access_token"] = access_token
            state["last_login_date"] = datetime.datetime.now(IST).strftime('%Y-%m-%d')
            save_state(state)
            logger.info("Successfully cached Zerodha access token in state.")
        else:
            send_telegram_message("❌ *Zerodha Login Failed*. Ensure the token is fresh and try again or send `login` to get a new link.")
    elif text.lower() == "login":
        trigger_zerodha_login(force=True)

# Background Telegram listener thread
def telegram_listener_thread():
    offset = 0
    try:
        r = requests.get(f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/getUpdates?offset=-1&timeout=0")
        if r.status_code == 200:
            res = r.json()
            if res.get("ok") and res.get("result"):
                offset = res["result"][-1]["update_id"] + 1
    except Exception as e:
        logger.error(f"Error initializing Telegram listener offset: {e}")

    logger.info("Telegram listener thread started...")
    
    while True:
        try:
            url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/getUpdates"
            params = {"offset": offset, "timeout": 10}
            r = requests.get(url, params=params, timeout=15)
            if r.status_code != 200:
                time.sleep(2)
                continue
                
            res = r.json()
            if not res.get("ok"):
                time.sleep(2)
                continue
                
            updates = res.get("result", [])
            for update in updates:
                offset = update["update_id"] + 1
                
                if "callback_query" in update:
                    handle_callback_query(update["callback_query"])
                elif "message" in update:
                    handle_text_message(update["message"])
                    
        except Exception as e:
            logger.error(f"Error in Telegram listener thread: {e}")
            time.sleep(5)

# Background Auto-Sell monitor thread running every 10 seconds during market hours
def autosell_monitor_thread():
    logger.info("Auto-sell monitor thread started...")
    while True:
        try:
            if is_market_hours():
                check_and_autosell_holdings()
            time.sleep(10)
        except Exception as e:
            logger.error(f"Error in auto-sell monitor thread: {e}")
            time.sleep(10)

def main():
    # Ensure single instance using a local socket lock
    import socket
    try:
        global _lock_socket
        _lock_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        _lock_socket.bind(('127.0.0.1', 48261))
    except socket.error:
        logger.warning("Another instance of MACD Crossover Scanner Bot is already running. Exiting.")
        sys.exit(0)

    logger.info("Initializing MACD Crossover Scanning Bot...")
    
    # Load state file first
    state = load_state()
    
    # Initialize ZerodhaClient and start Telegram listener thread
    global zerodha_client
    api_key = os.getenv("ZERODHA_API_KEY")
    api_secret = os.getenv("ZERODHA_API_SECRET")
    client_id = os.getenv("ZERODHA_CLIENT_ID")
    
    if all([api_key, api_secret, client_id]):
        logger.info("Initializing Zerodha client...")
        zerodha_client = ZerodhaClient(api_key, api_secret, client_id)
        zerodha_client.initialize_client()
        
        # Start Telegram Listener Thread
        listener = threading.Thread(target=telegram_listener_thread, daemon=True)
        listener.start()
        
        # Start Auto-Sell Checker Thread
        autosell_thread = threading.Thread(target=autosell_monitor_thread, daemon=True)
        autosell_thread.start()
        
        # Send Bot Start Notification
        send_telegram_message("🤖 *MACD Crossover Scanner Bot Started Online*")
        
        # Check if we have a saved token in state
        saved_token = state.get("zerodha_access_token")
        session_restored = False
        if saved_token:
            logger.info("Found saved Zerodha access token. Verifying session...")
            zerodha_client.kite.set_access_token(saved_token)
            if zerodha_client.validate_session():
                logger.info("Zerodha session successfully restored from saved token!")
                send_telegram_message("✅ *Zerodha Session Restored*: Connected successfully.")
                session_restored = True
            else:
                logger.warning("Saved Zerodha access token is invalid or expired.")
                
        if not session_restored:
            # Trigger login to request token
            trigger_zerodha_login(force=False)
    else:
        logger.warning("Zerodha credentials incomplete. Order placement is disabled.")
        send_telegram_message("🤖 *MACD Crossover Scanner Bot Started Online* (Zerodha Disabled)")
    
    while True:
        try:
            now = datetime.datetime.now(IST)
            today_str = now.strftime('%Y-%m-%d')
            
            # Trigger Daily Summary at or after 3:45 PM IST on weekdays
            if now.weekday() < 5 and now.time() >= datetime.time(15, 45):
                state = load_state()
                if state.get("last_summary_date") != today_str:
                    logger.info("Triggering Daily Summary report...")
                    send_daily_summary(state)
            
            if is_market_hours():
                logger.info("Market is open. Running scan cycle...")
                state = load_state()
                run_scan_cycle(state)
                # Sleep for 5 minutes (300 seconds) before the next scan cycle
                logger.info("Scan completed. Sleeping for 5 minutes...")
                time.sleep(300)
            else:
                # Calculate sleep duration to keep logs clean - check every 1 minute
                logger.debug("Market is closed. Idle monitoring...")
                time.sleep(60)
        except KeyboardInterrupt:
            logger.info("Bot manually terminated by user.")
            send_telegram_message("🛑 *MACD Crossover Scanner Bot Stopped*")
            break
        except Exception as e:
            logger.error(f"Error in main loop: {e}")
            send_telegram_message(f"⚠️ *Error in MACD Scanner Bot*: `{e}`")
            time.sleep(60) # wait before retry

if __name__ == "__main__":
    main()
